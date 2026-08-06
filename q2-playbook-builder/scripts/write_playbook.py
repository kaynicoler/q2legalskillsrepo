#!/usr/bin/env python3
"""Populate the Q2 playbook template with rules + summary sheets.
 
Usage:
    python3 write_playbook.py \
        --template references/playbook-template.xlsx \
        --rules rules.json \
        --exec-summary exec_summary.md \
        --open-issues open_issues.md \
        --ai-prompts ai_prompts.md \
        --clause-library clause_library.md \
        --out "Q2 [Contract Type] - Contract Review Playbook.xlsx"
 
The rules JSON file must be a list of objects with these 17 keys:
    name, type, text, fallback_1, fallback_2, fallback_3,
    reviewer_note, preferred_language_1, preferred_language_2, preferred_language_3,
    suggested_comment_1, suggested_comment_2, suggested_comment_3,
    recommendation_prompt, low_risk, medium_risk, high_risk
 
All summary inputs are markdown files. Each top-level heading becomes a section header
in the corresponding Excel sheet; everything below is the section body.
"""
import argparse
import json
import shutil
import os
import sys
import re
from pathlib import Path
 
try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)
 
 
# Order matters — must match the 17-column playbook schema
RULE_KEYS = [
    'name', 'type', 'text',
    'fallback_1', 'fallback_2', 'fallback_3',
    'reviewer_note',
    'preferred_language_1', 'preferred_language_2', 'preferred_language_3',
    'suggested_comment_1', 'suggested_comment_2', 'suggested_comment_3',
    'recommendation_prompt',
    'low_risk', 'medium_risk', 'high_risk',
]
 
VALID_TYPES = {'Rule', 'Question', 'Risk'}
 
 
def validate_rules(rules):
    errors = []
    for i, r in enumerate(rules, 1):
        missing = [k for k in RULE_KEYS if k not in r]
        if missing:
            errors.append(f"Rule {i}: missing keys {missing}")
        if r.get('type') not in VALID_TYPES:
            errors.append(f"Rule {i}: type must be one of {VALID_TYPES}, got {r.get('type')!r}")
    if errors:
        for e in errors:
            print(f"  {e}", file=sys.stderr)
        sys.exit(1)
 
 
def parse_markdown_sections(md_text):
    """Parse a markdown file into [(header, body)] tuples.
 
    Top-level # headings become section headers. Everything up to the next
    top-level heading is the body. If there are no top-level headings, the
    entire content becomes a single section keyed by the filename.
    """
    lines = md_text.splitlines()
    sections = []
    current_header = None
    current_body = []
 
    for line in lines:
        if line.startswith('# '):
            if current_header is not None:
                sections.append((current_header, '\n'.join(current_body).strip()))
            current_header = line[2:].strip()
            current_body = []
        else:
            current_body.append(line)
 
    if current_header is not None:
        sections.append((current_header, '\n'.join(current_body).strip()))
    elif current_body:
        sections.append(('Content', '\n'.join(current_body).strip()))
 
    return sections
 
 
def add_text_sheet(wb, title, content_blocks):
    sh = wb.create_sheet(title)
    sh.column_dimensions['A'].width = 110
    row = 1
    for header, body in content_blocks:
        h = sh.cell(row=row, column=1, value=header)
        h.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', start_color='1F4E79')
        h.alignment = Alignment(wrap_text=True, vertical='center')
        sh.row_dimensions[row].height = 25
        row += 1
        c = sh.cell(row=row, column=1, value=body)
        c.font = Font(name='Arial', size=10)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        line_count = body.count('\n') + 1
        sh.row_dimensions[row].height = max(40, min(600, line_count * 16 + 40))
        row += 2
 
 
def write_playbook(template, rules, summaries, out_path):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(template, out_path)
    try:
        os.chmod(out_path, 0o644)
    except (PermissionError, OSError):
        pass
 
    wb = load_workbook(out_path)
    if 'PlaybookChecks' not in wb.sheetnames:
        print(f"Template missing 'PlaybookChecks' sheet", file=sys.stderr)
        sys.exit(1)
    ws = wb['PlaybookChecks']
 
    # Write rules
    for i, rule in enumerate(rules, start=2):
        for j, key in enumerate(RULE_KEYS, start=1):
            v = rule.get(key, '')
            cell = ws.cell(row=i, column=j, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Arial', size=10)
 
    # Format header
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col)
        h.font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        h.fill = PatternFill('solid', start_color='1F4E79')
        h.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
 
    # Column widths
    widths = {1: 28, 2: 10, 3: 60, 4: 50, 5: 50, 6: 50, 7: 60, 8: 60, 9: 60, 10: 60,
              11: 50, 12: 50, 13: 50, 14: 60, 15: 35, 16: 35, 17: 35}
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w
 
    # Row heights
    ws.row_dimensions[1].height = 40
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 250
 
    ws.freeze_panes = 'A2'
 
    # Add summary sheets in this order
    sheet_order = ['Executive Summary', 'Open Issues', 'AI Review Prompts', 'Clause Library']
    for title in sheet_order:
        blocks = summaries.get(title)
        if blocks:
            add_text_sheet(wb, title, blocks)
 
    # Reorder sheets: PlaybookChecks, summaries, CheckTypes
    final_order = ['PlaybookChecks'] + sheet_order + ['CheckTypes']
    wb._sheets = [wb[n] for n in final_order if n in wb.sheetnames]
 
    wb.save(out_path)
    print(f"Saved {out_path}")
    print(f"Playbook rows: {len(rules)}")
 
 
def main():
    ap = argparse.ArgumentParser(description="Populate the Q2 playbook template.")
    ap.add_argument('--template', required=True, help='Path to playbook-template.xlsx')
    ap.add_argument('--rules', required=True, help='Path to rules JSON file')
    ap.add_argument('--exec-summary', help='Path to Executive Summary markdown')
    ap.add_argument('--open-issues', help='Path to Open Issues markdown')
    ap.add_argument('--ai-prompts', help='Path to AI Review Prompts markdown')
    ap.add_argument('--clause-library', help='Path to Clause Library markdown')
    ap.add_argument('--out', required=True, help='Output xlsx path')
    args = ap.parse_args()
 
    with open(args.rules) as f:
        rules = json.load(f)
    validate_rules(rules)
 
    summaries = {}
    summary_args = {
        'Executive Summary': args.exec_summary,
        'Open Issues': args.open_issues,
        'AI Review Prompts': args.ai_prompts,
        'Clause Library': args.clause_library,
    }
    for sheet_name, path in summary_args.items():
        if path and os.path.exists(path):
            with open(path) as f:
                summaries[sheet_name] = parse_markdown_sections(f.read())
 
    write_playbook(args.template, rules, summaries, args.out)
 
 
if __name__ == '__main__':
    main()
